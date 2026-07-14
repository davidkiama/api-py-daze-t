import pandas as pd
from datetime import datetime, timezone
import calendar
import pytz
import oandapyV20
from oandapyV20.contrib.factories import InstrumentsCandlesFactory
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from dotenv import load_dotenv


# Load Env
load_dotenv()

# Initialize OANDA Client outside the function so it can be reused across requests
OANDA_ACCESS_TOKEN = os.getenv("OANDA_API_TOKEN", "YOUR_OANDA_TOKEN")
ENVIRONMENT = "practice"
client = oandapyV20.API(access_token=OANDA_ACCESS_TOKEN,
                        environment=ENVIRONMENT)


def fetch_oanda_data(instrument, granularity, from_time, to_time):
    params = {
        "from": from_time,
        "to": to_time,
        "granularity": granularity,
        "price": "M"
    }

    data = []
    try:
        for r in InstrumentsCandlesFactory(instrument=instrument, params=params):
            client.request(r)
            for c in r.response.get('candles', []):
                if c['complete']:
                    data.append({
                        'time': c['time'],
                        'open': float(c['mid']['o']),
                        'high': float(c['mid']['h']),
                        'low': float(c['mid']['l']),
                        'close': float(c['mid']['c']),
                    })

        df = pd.DataFrame(data)
        if not df.empty:
            df['time'] = pd.to_datetime(df['time'])
        return df
    except Exception as e:
        print(f"Error fetching data for {instrument}: {e}")
        return pd.DataFrame()


def analyze_day_strategy(day_df, instrument_name):
    ny_tz = pytz.timezone('America/New_York')
    day_df['time_ny'] = day_df['time'].dt.tz_convert(ny_tz)
    day_df = day_df.sort_values('time_ny').reset_index(drop=True)

    date_str = day_df['time_ny'].iloc[0].strftime('%Y-%m-%d')

    # Time filtering in minutes
    time_mins = day_df['time_ny'].dt.hour * 60 + day_df['time_ny'].dt.minute

    session_window = day_df[(time_mins >= 0) & (
        time_mins < 570)]       # 00:00 - 09:30
    open_window = day_df[(time_mins >= 570) & (
        time_mins < 600)]        # 09:30 - 10:00
    reversal_window = day_df[(time_mins >= 600) & (
        time_mins <= 660)]   # 10:00 - 11:00

    if session_window.empty or open_window.empty or reversal_window.empty:
        return None

    # 1. Structural Values
    idx_sh = session_window['high'].idxmax()
    idx_sl = session_window['low'].idxmin()
    idx_oh = open_window['high'].idxmax()
    idx_ol = open_window['low'].idxmin()

    session_h = session_window.loc[idx_sh, 'high']
    session_l = session_window.loc[idx_sl, 'low']
    open_h = open_window.loc[idx_oh, 'high']
    open_l = open_window.loc[idx_ol, 'low']

    # 2. Run Bias calculation
    run_bias = "Neutral"
    if open_h >= session_h:
        run_bias = "Bullish"
    elif open_l <= session_l:
        run_bias = "Bearish"

    # 3. Reversal Window tracking
    day_high = max(session_h, open_h)
    day_low = min(session_l, open_l)

    idx_rev_h = reversal_window['high'].idxmax()
    idx_rev_l = reversal_window['low'].idxmin()
    rev_high = reversal_window.loc[idx_rev_h, 'high']
    rev_low = reversal_window.loc[idx_rev_l, 'low']

    new_high_formed = rev_high > day_high
    new_low_formed = rev_low < day_low

    new_high_low_status = "No"
    new_hl_time = "N/A"
    if new_high_formed:
        new_high_low_status = "Yes (High)"
        new_hl_time = reversal_window.loc[idx_rev_h, 'time_ny'].strftime(
            '%H:%M')
    elif new_low_formed:
        new_high_low_status = "Yes (Low)"
        new_hl_time = reversal_window.loc[idx_rev_l, 'time_ny'].strftime(
            '%H:%M')

    # 4. Fib 0.236 Retracement Check
    fib_touched = "No"
    fib_time = "N/A"

    if new_high_formed:
        peak_time = reversal_window.loc[idx_rev_h, 'time_ny']
        post_peak_df = reversal_window[reversal_window['time_ny'] > peak_time]
        fib_target = rev_high - ((rev_high - day_low) * 0.236)

        touches = post_peak_df[post_peak_df['low'] <= fib_target]
        if len(touches) > 0:
            fib_touched = "Yes"
            fib_time = touches.iloc[0]['time_ny'].strftime('%H:%M')

    elif new_low_formed:
        trough_time = reversal_window.loc[idx_rev_l, 'time_ny']
        post_trough_df = reversal_window[reversal_window['time_ny'] > trough_time]
        fib_target = rev_low + ((day_high - rev_low) * 0.236)

        touches = post_trough_df[post_trough_df['high'] >= fib_target]
        if len(touches) > 0:
            fib_touched = "Yes"
            fib_time = touches.iloc[0]['time_ny'].strftime('%H:%M')

    return {
        "Date": date_str,
        "Pair": instrument_name,
        "Session-High": session_h,
        "Session-High-Time": session_window.loc[idx_sh, 'time_ny'].strftime('%H:%M'),
        "Session-Low": session_l,
        "Session-Low-Time": session_window.loc[idx_sl, 'time_ny'].strftime('%H:%M'),
        "Run Bias": run_bias,
        "New High/Low Made (Yes/No)": new_high_low_status,
        "New-High-Low Time": new_hl_time,
        "Price Retraced & Touched 0.236": fib_touched,
        "Time Touched 0.236": fib_time
    }


def generate_excel_report(data_rows, filename):
    df = pd.DataFrame(data_rows)

    # Sort by Date then Pair to keep the output incredibly clean
    df = df.sort_values(by=["Date", "Pair"]).reset_index(drop=True)

    # Inject the manual tracking columns
    df["Traded? (Manual Yes/No)"] = ""
    df["Execution Bias Taken (Manual)"] = ""
    df["Risk to Reward Achieved (Manual)"] = ""
    df["Trade Notes (Manual)"] = ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1-Month Backtest Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Spreadsheet Formatting
    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    manual_fill = PatternFill(start_color="2E75B6",
                              end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        cell.fill = manual_fill if "Manual" in str(cell.value) else header_fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max([len(str(c.value))
                      for c in col if c.value is not None] + [0])
        ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

    ws.freeze_panes = "A2"
    wb.save(filename)


def run_backtest_for_month(instrument, year, month):
    """
    Called by Flask endpoint. Runs the full backtest logic for the requested month,
    generates the formatted Excel file, and returns the filepath.
    """
    granularity = "M5"
    _, last_day = calendar.monthrange(year, month)

    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    end_date = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

    from_time = start_date.strftime('%Y-%m-%dT%H:%M:%SZ')
    to_time = end_date.strftime('%Y-%m-%dT%H:%M:%SZ')

    print(f"Pulling data for {instrument} for {month}/{year}...")
    raw_df = fetch_oanda_data(instrument, granularity, from_time, to_time)

    all_extracted_rows = []

    if not raw_df.empty:
        # Generate a unified date column to group by day
        ny_tz = pytz.timezone('America/New_York')
        raw_df['date_group'] = raw_df['time'].dt.tz_convert(
            ny_tz).dt.strftime('%Y-%m-%d')

        for day, group_df in raw_df.groupby('date_group'):
            result = analyze_day_strategy(group_df.copy(), instrument)
            if result:
                all_extracted_rows.append(result)

    # Generate filename (saved in current working directory)
    filename = f"datasets/backtest_{instrument}_{year}_{month:02d}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)

    if all_extracted_rows:
        generate_excel_report(all_extracted_rows, filepath)
    else:
        # If absolutely no trades occurred that month, still generate a blank structured file
        generate_excel_report([], filepath)

    return filepath
